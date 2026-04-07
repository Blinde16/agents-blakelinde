"""User-scoped spreadsheet staging (CSV/XLSX) and CFO query tools (asyncpg)."""

from __future__ import annotations

import csv
import io
import json
import re
from typing import Any
from uuid import uuid4

from src.orchestration.db_pool import get_pool
from src.tools.schemas import QueryStagingMetricsInput, SummarizeSheetInput

MAX_INGEST_ROWS = 10_000
SUMMARY_SAMPLE_ROWS = 5_000


def _norm_header(label: str) -> str:
    raw = (label or "").strip().lower()
    raw = re.sub(r"\s+", "_", raw)
    raw = re.sub(r"[^a-z0-9_]", "", raw)
    return raw or "column"


def _unique_headers(labels: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for lab in labels:
        base = _norm_header(lab)
        n = seen.get(base, 0)
        seen[base] = n + 1
        out.append(base if n == 0 else f"{base}_{n + 1}")
    return out


def _parse_csv_bytes(data: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = data.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows_list = list(reader)
    if not rows_list:
        return [], []
    headers = _unique_headers([str(h) for h in rows_list[0]])
    out: list[dict[str, Any]] = []
    for raw in rows_list[1:]:
        if not raw or all((c or "").strip() == "" for c in raw):
            continue
        padded = list(raw) + [""] * max(0, len(headers) - len(raw))
        row_obj = {headers[i]: (padded[i] if i < len(padded) else "") for i in range(len(headers))}
        out.append(row_obj)
    return headers, out[:MAX_INGEST_ROWS]


def _parse_xlsx_bytes(data: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    from openpyxl import load_workbook  # lazy import

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        first = next(rows_iter, None)
        if first is None:
            return [], []
        headers = _unique_headers([str(h) if h is not None else "" for h in first])
        out: list[dict[str, Any]] = []
        for raw in rows_iter:
            if raw is None:
                continue
            cells = list(raw)
            if not cells or all(v is None or str(v).strip() == "" for v in cells):
                continue
            padded = list(cells) + [None] * max(0, len(headers) - len(cells))
            row_obj = {
                headers[i]: ("" if padded[i] is None else str(padded[i]).strip())
                for i in range(len(headers))
            }
            out.append(row_obj)
            if len(out) >= MAX_INGEST_ROWS:
                break
        return headers, out
    finally:
        wb.close()


def parse_spreadsheet(filename: str, data: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return _parse_csv_bytes(data)
    if lower.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx_bytes(data)
    raise ValueError("Unsupported file type. Use .csv, .xlsx, or .xlsm.")


async def ingest_user_spreadsheet(
    user_internal_id: str,
    filename: str,
    file_bytes: bytes,
) -> dict[str, Any]:
    columns, rows = parse_spreadsheet(filename, file_bytes)
    upload_id = uuid4()
    uid = user_internal_id
    fn = filename

    if not rows:
        raise ValueError("No data rows after header.")

    batch: list[tuple[str, str, str, int, str]] = []
    async with get_pool().acquire() as conn:
        async with conn.transaction():
            for i, row in enumerate(rows, start=1):
                batch.append(
                    (
                        uid,
                        str(upload_id),
                        fn,
                        i,
                        json.dumps(row, ensure_ascii=False),
                    )
                )
                if len(batch) >= 400:
                    await conn.executemany(
                        """
                        INSERT INTO public.finance_sheet_staging
                            (user_id, upload_id, filename, row_index, row_payload)
                        VALUES ($1::uuid, $2::uuid, $3, $4, $5::jsonb)
                        """,
                        batch,
                    )
                    batch = []
            if batch:
                await conn.executemany(
                    """
                    INSERT INTO public.finance_sheet_staging
                        (user_id, upload_id, filename, row_index, row_payload)
                    VALUES ($1::uuid, $2::uuid, $3, $4, $5::jsonb)
                    """,
                    batch,
                )

    return {
        "upload_id": str(upload_id),
        "filename": fn,
        "columns": columns,
        "rows_ingested": len(rows),
    }


async def query_staging_metrics(
    upload_id: str,
    user_internal_id: str,
    limit: int,
    offset: int,
) -> str:
    try:
        parsed = QueryStagingMetricsInput(upload_id=upload_id, limit=limit, offset=offset)
    except Exception as exc:  # noqa: BLE001
        return json.dumps({"error": "validation_error", "detail": str(exc)})

    uid = user_internal_id
    try:
        async with get_pool().acquire() as conn:
            check = await conn.fetchrow(
                """
                SELECT 1 AS ok
                FROM public.finance_sheet_staging
                WHERE user_id = $1::uuid AND upload_id = $2::uuid AND row_index > 0
                LIMIT 1
                """,
                uid,
                str(parsed.upload_id),
            )
            if check is None:
                return json.dumps(
                    {"detail": "No staged rows for this upload_id, or access denied."}
                )

            recs = await conn.fetch(
                """
                SELECT row_index, row_payload
                FROM public.finance_sheet_staging
                WHERE user_id = $1::uuid AND upload_id = $2::uuid AND row_index > 0
                ORDER BY row_index
                LIMIT $3 OFFSET $4
                """,
                uid,
                str(parsed.upload_id),
                parsed.limit,
                parsed.offset,
            )
    except Exception as exc:  # noqa: BLE001
        return json.dumps({"error": "database_error", "detail": str(exc)})

    rows_out = [{"row_index": r["row_index"], **dict(r["row_payload"])} for r in recs]
    return json.dumps(
        {
            "upload_id": str(parsed.upload_id),
            "limit": parsed.limit,
            "offset": parsed.offset,
            "rows": rows_out,
        },
        ensure_ascii=False,
    )


def _numeric_stats(values: list[str]) -> dict[str, Any] | None:
    nums: list[float] = []
    for v in values:
        if v is None or str(v).strip() == "":
            continue
        try:
            nums.append(float(str(v).replace(",", "")))
        except ValueError:
            return None
    if not nums:
        return None
    return {
        "count": len(nums),
        "min": min(nums),
        "max": max(nums),
        "sum": sum(nums),
        "mean": sum(nums) / len(nums),
    }


async def summarize_sheet(upload_id: str, user_internal_id: str) -> str:
    try:
        parsed = SummarizeSheetInput(upload_id=upload_id)
    except Exception as exc:  # noqa: BLE001
        return json.dumps({"error": "validation_error", "detail": str(exc)})

    uid = user_internal_id
    uid_str = str(parsed.upload_id)

    try:
        async with get_pool().acquire() as conn:
            total = await conn.fetchval(
                """
                SELECT COUNT(*)::int
                FROM public.finance_sheet_staging
                WHERE user_id = $1::uuid AND upload_id = $2::uuid
                  AND row_index > 0
                """,
                uid,
                uid_str,
            )
            if total == 0:
                return json.dumps(
                    {"detail": "No staged rows for this upload_id, or access denied."}
                )

            recs = await conn.fetch(
                """
                SELECT row_payload
                FROM public.finance_sheet_staging
                WHERE user_id = $1::uuid AND upload_id = $2::uuid AND row_index > 0
                ORDER BY row_index
                LIMIT $3
                """,
                uid,
                uid_str,
                SUMMARY_SAMPLE_ROWS,
            )
    except Exception as exc:  # noqa: BLE001
        return json.dumps({"error": "database_error", "detail": str(exc)})

    payloads = [dict(r["row_payload"]) for r in recs]
    all_keys: list[str] = []
    seen_k: set[str] = set()
    for p in payloads:
        for k in p:
            if k not in seen_k:
                seen_k.add(k)
                all_keys.append(k)

    column_stats: dict[str, Any] = {}
    for key in all_keys:
        vals = [p.get(key, "") for p in payloads]
        str_vals = [str(v) if v is not None else "" for v in vals]
        stats = _numeric_stats(str_vals)
        if stats is not None:
            column_stats[key] = stats

    sample = payloads[:5]
    sample_truncated = total > SUMMARY_SAMPLE_ROWS

    return json.dumps(
        {
            "upload_id": uid_str,
            "row_count": total,
            "columns": all_keys,
            "numeric_column_stats": column_stats,
            "sample_rows": sample,
            "sample_truncated": sample_truncated,
        },
        ensure_ascii=False,
    )
